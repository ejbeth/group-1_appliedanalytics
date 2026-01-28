from flask import Flask, render_template, redirect, url_for, session, flash, request, jsonify
import os
from openpyxl import load_workbook
from datetime import datetime, timedelta

app = Flask(__name__)
app.secret_key = 'amefa-portal-secret-key-2024'  # Change this for production

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "amefa_order_portal_demo_db_final_en.xlsx")

SHEET_ORDERS = "Orders"                                # your sheet name
SHEET_CUSTOMERS = "Customers"  # optional (only if you want per-customer filtering)
SHEET_SUPPORT = "Support"

# Ensure that the Support worksheet exists in the Excel database.
# If the sheet does not exist, it is created with the required header structure.
# This helper guarantees a consistent schema before writing support ticket data.
def ensure_support_sheet(wb):
    if SHEET_SUPPORT not in wb.sheetnames:
        ws = wb.create_sheet(SHEET_SUPPORT)
        ws.append([
            "TicketID",
            "CustomerID",
            "OrderID",
            "EmailTracking",
            "CreatedAt",
            "Status"
        ])
        return ws
    return wb[SHEET_SUPPORT]

# Debug: Show template structure
print("=" * 50)
print("AMEFA ORDER PORTAL - Flask Application")
print("=" * 50)
print(f"Current directory: {os.getcwd()}")
print(f"Templates folder: {app.template_folder}")
print("=" * 50)

# Root route - redirects to login
@app.route('/')
def home():
    # Redirect to login page
    return redirect(url_for('login'))

# Login route (using your cust_login_reg.html template)
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        # Simple authentication (replace with real auth later)
        username = request.form.get('username')
        password = request.form.get('password')
        
        if username and password:  # Basic validation
            session['user_id'] = 1  # Set a dummy user ID
            session['username'] = username
            session['customer_id'] = resolve_customer_id_from_username(username)
            flash('Login successful! Welcome to Amefa Order Portal.', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Please enter both username and password', 'error')
    
    # Render the login page
    return render_template('pages/cust_login_reg.html')

# Logout route
@app.route('/logout')
def logout():
    session.clear()
    flash('You have been logged out successfully.', 'info')
    return redirect(url_for('login'))

def normalize(s):
    return (str(s).strip().lower() if s is not None else "")

def parse_eur(value):
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)

    s = str(value).strip()
    s = s.replace("€", "").replace("EUR", "").replace(" ", "")
    s = s.replace(".", "").replace(",", ".")

    try:
        return float(s)
    except Exception:
        return 0.0

def resolve_customer_id_from_username(username: str):
    """
    Map session username -> CustomerID using Customers sheet.
    Tries matching against: CustomerID, CustomerName, Email (case-insensitive).
    Returns customer_id or None.
    """
    if not username:
        return None

    u = normalize(username)

    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET_CUSTOMERS]

    # header map
    headers = {}
    for col in range(1, ws.max_column + 1):
        name = ws.cell(row=1, column=col).value
        if name:
            headers[normalize(name)] = col

    def col(name):
        return headers.get(normalize(name))

    c_id = col("CustomerID")
    c_name = col("CustomerName")
    c_email = col("Email")

    customer_id = None
    for r in range(2, ws.max_row + 1):
        cid = ws.cell(r, c_id).value if c_id else None
        cname = ws.cell(r, c_name).value if c_name else None
        cemail = ws.cell(r, c_email).value if c_email else None

        if u in (normalize(cid), normalize(cname), normalize(cemail)):
            customer_id = str(cid)
            break

    wb.close()
    return customer_id

# Dashboard route
def load_metrics_from_excel(customer_id=None):
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET_ORDERS]

    # build header map (row 1)
    headers = {}
    for col in range(1, ws.max_column + 1):
        name = ws.cell(row=1, column=col).value
        if name:
            headers[str(name).strip().lower()] = col

    def col(name):
        return headers.get(name.lower())

    c_order_id = col("orderid")
    c_date = col("orderdate")
    c_customer = col("customerid")
    c_status = col("status")
    c_total = col("total")
    c_delivery = col("expecteddelivery")

    orders = []
    for r in range(2, ws.max_row + 1):
        oid = ws.cell(r, c_order_id).value if c_order_id else None
        if not oid:
            continue

        cust = ws.cell(r, c_customer).value if c_customer else None
        if customer_id and normalize(cust) != normalize(customer_id):
            continue

        date_val = ws.cell(r, c_date).value if c_date else None
        status = ws.cell(r, c_status).value if c_status else "Unknown"
        total_val = ws.cell(r, c_total).value if c_total else 0
        deliv_val = ws.cell(r, c_delivery).value if c_delivery else None

        # normalize money
        total_num = parse_eur(total_val)

        # normalize date
        if isinstance(date_val, datetime):
            date_str = date_val.strftime("%Y-%m-%d")
        else:
            date_str = str(date_val) if date_val else ""

        orders.append({
            "order_id": str(oid).replace("#", ""),
            "date_raw": date_val,
            "date": date_str,
            "status": str(status),
            "total_value": total_num,
            "total": f"€ {total_num:,.2f}",
            "delivery_raw": deliv_val,
        })

    # metrics
    today = datetime.today()
    last_30 = today - timedelta(days=30)
    last_90 = today - timedelta(days=90)

    def to_dt(v):
        if isinstance(v, datetime):
            return v
        return None

    orders_30d = 0
    spent_90d = 0.0
    status_counts = {}
    open_orders = 0
    next_delivery = None

    for o in orders:
        dt = to_dt(o["date_raw"])
        if dt and dt >= last_30:
            orders_30d += 1
        if dt and dt >= last_90:
            spent_90d += o["total_value"]

        status_counts[o["status"]] = status_counts.get(o["status"], 0) + 1
        if o["status"] in ("Processing", "Pending"):
            open_orders += 1

        ddt = to_dt(o["delivery_raw"])
        if ddt and (next_delivery is None or ddt < next_delivery):
            next_delivery = ddt

    recent_orders = sorted(
        orders,
        key=lambda x: (to_dt(x["date_raw"]) or datetime.min),
        reverse=True
    )[:3]

    metrics = {
        "orders_30d": orders_30d,
        "orders_30d_delta": None,
        "open_orders": open_orders,
        "spent_90d": f"€ {spent_90d:,.2f}",
        "next_delivery": next_delivery.strftime("%b %d") if next_delivery else "—",
        "status_counts": status_counts,
        "recent_orders": recent_orders,
    }
    # --- Support tickets (latest 5) ---
    support_tickets = []
    if SHEET_SUPPORT in wb.sheetnames:
        ws_sup = wb[SHEET_SUPPORT]

        for row in ws_sup.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue

            # Columns: TicketID, CustomerID, OrderID, EmailTracking, CreatedAt, Status
            t_customer = row[1]

            # Filter by logged-in customer
            if customer_id and normalize(t_customer) != normalize(customer_id):
                continue

            support_tickets.append({
                "ticket_id": row[0],
                "customer_id": row[1],
                "order_id": row[2],
                "email_tracking": row[3],
                "created_at": row[4],
                "status": row[5],
            })

        # Sort by CreatedAt (string "YYYY-MM-DD HH:MM:SS") and take latest 5
        support_tickets = sorted(
            support_tickets,
            key=lambda x: x.get("created_at") or "",
            reverse=True
        )[:5]

    metrics["recent_tickets"] = support_tickets
    metrics["open_tickets"] = sum(1 for t in support_tickets if (t.get("status") or "").lower() == "open")

    wb.close()
    return metrics

def load_orders_from_excel(customer_id=None):
    """
    Load all orders from the Excel database.
    If customer_id is provided, filter results to the logged-in customer.
    Returns a list of order dictionaries for the Orders page.
    """
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET_ORDERS]

    # Build a header -> column index map from the first row
    headers = {}
    for c in range(1, ws.max_column + 1):
        name = ws.cell(row=1, column=c).value
        if name:
            headers[str(name).strip().lower()] = c

    def col(name):
        return headers.get(name.lower())

    c_order_id = col("orderid")
    c_date = col("orderdate")
    c_customer = col("customerid")
    c_status = col("status")
    c_total = col("total")

    orders = []

    # Read all rows and optionally filter by customer_id
    for r in range(2, ws.max_row + 1):
        oid = ws.cell(r, c_order_id).value if c_order_id else None
        if not oid:
            continue

        cust = ws.cell(r, c_customer).value if c_customer else None
        if customer_id and normalize(cust) != normalize(customer_id):
            continue

        date_val = ws.cell(r, c_date).value if c_date else None
        status_val = ws.cell(r, c_status).value if c_status else "Unknown"
        total_val = ws.cell(r, c_total).value if c_total else 0

        # Normalize values for UI rendering
        total_num = parse_eur(total_val)
        date_str = date_val.strftime("%Y-%m-%d") if isinstance(date_val, datetime) else ""

        orders.append({
            "order_id": str(oid).replace("#", ""),
            "date": date_str,
            "status": str(status_val),
            "total_value": total_num,
            "total": f"€ {total_num:,.2f}",
        })

    # Sort newest first
    orders = sorted(orders, key=lambda x: x.get("date") or "", reverse=True)

    wb.close()
    return orders
    
@app.route('/dashboard')
def dashboard():
    if 'user_id' not in session:
        flash('Please login to access the dashboard.', 'error')
        return redirect(url_for('login'))

    username = session.get('username')
    customer_id = resolve_customer_id_from_username(username)

    metrics = load_metrics_from_excel(customer_id=customer_id)

    return render_template(
        'pages/dashboard.html',
        username=session.get('username'),
        metrics=metrics
    )

# Products route
@app.route('/products')
def products():
    if 'user_id' not in session:
        flash('Please login to view products.', 'error')
        return redirect(url_for('login'))
    return render_template('pages/products.html')

# Orders route
@app.route('/orders')
def orders():
    if 'user_id' not in session:
        flash('Please login to view orders.', 'error')
        return redirect(url_for('login'))

    # Get customer linked to the current session
    customer_id = session.get("customer_id")

    # Load all orders for this customer
    orders = load_orders_from_excel(customer_id=customer_id)

    return render_template(
        'pages/order_view.html',
        orders=orders
    )

# Reports route
@app.route('/reports')
def reports():
    if 'user_id' not in session:
        flash('Please login to view reports.', 'error')
        return redirect(url_for('login'))
    return render_template('pages/report.html')

# Support route
@app.route('/support')
def support():
    if 'user_id' not in session:
        flash('Please login to access support.', 'error')
        return redirect(url_for('login'))
    return render_template('pages/support.html')

# Create a support ticket and store it in the Excel database
@app.route("/support/create", methods=["POST"])
def support_create():
    # Read JSON payload from request
    data = request.get_json(silent=True) or {}

    # Extract required fields
    ticket_id = (data.get("ticket_id") or "").strip()
    order_id = (data.get("order_id") or "").strip()
    email_tracking = (data.get("email_tracking") or "").strip()

    # Get customer from session (if logged in)
    customer_id = session.get("customer_id")

    # Set metadata
    created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    status = "Open"

    # Basic validation
    if not ticket_id:
        return jsonify({"ok": False, "error": "Missing ticket_id"}), 400
    if not email_tracking:
        return jsonify({"ok": False, "error": "Missing email_tracking"}), 400

    # Write ticket to Excel
    wb = load_workbook(EXCEL_PATH)
    ws = ensure_support_sheet(wb)
    ws.append([ticket_id, customer_id, order_id, email_tracking, created_at, status])
    wb.save(EXCEL_PATH)
    wb.close()

    # Return success response
    return jsonify({"ok": True})

# Add a test route to verify CSS is working
@app.route('/test-css')
def test_css():
    """Test route to verify CSS is loading correctly"""
    return render_template('pages/dashboard.html', username='Test User')

if __name__ == '__main__':
    print("\n" + "=" * 50)
    print("🚀 AMEFA ORDER PORTAL STARTING...")
    print("=" * 50)
    print("Application is running!")
    print("Open your browser and go to: http://localhost:5000")
    print("=" * 50)
    print("TEST CREDENTIALS:")
    print("• Username: any username")
    print("• Password: any password")
    print("=" * 50)
    print("Navigation:")
    print("• /login - Login page")
    print("• /dashboard - Main dashboard (after login)")
    print("• /products - Products catalog")
    print("• /orders - Order management")
    print("• /reports - Reports & analytics")
    print("• /support - Support center")
    print("• /logout - Logout")
    print("=" * 50 + "\n")
    
    app.run(debug=True, port=5000)